from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, make_response
from flask_login import (
    LoginManager, UserMixin,
    login_user, logout_user, login_required, current_user,
)
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from pathlib import Path
import sqlite3
import csv as csv_mod
import io
import os

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "linkbot-dev-secret")

LINKS_FILE = Path(os.environ.get("LINKS_FILE", "links.txt"))
ENV_FILE   = Path(os.environ.get("ENV_FILE",   ".env"))
DB_FILE    = Path(os.environ.get("DB_FILE",    "db.sqlite3"))
APP_PORT   = int(os.environ.get("APP_PORT", 5000))
APP_URL    = os.environ.get("APP_URL", "http://localhost")

# ── Template filtre ───────────────────────────────────────────────────────────

_AVATAR_COLORS = ["#4f46e5","#0891b2","#059669","#d97706","#dc2626","#7c3aed","#db2777"]

@app.template_filter("avatar_color")
def avatar_color(username: str) -> str:
    if not username:
        return "#6b7280"
    return _AVATAR_COLORS[ord(username[0].lower()) % len(_AVATAR_COLORS)]


@app.template_filter("fdate")
def fdate(value: str) -> str:
    if not value:
        return "—"
    try:
        dt = datetime.strptime(str(value), "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return str(value)[:16]


# ── Flask-Login ───────────────────────────────────────────────────────────────

login_manager = LoginManager(app)
login_manager.login_view             = "login"
login_manager.login_message          = "Bu sayfaya erişmek için giriş yapmalısınız."
login_manager.login_message_category = "warning"


class User(UserMixin):
    def __init__(self, id: int, username: str, email: str = ""):
        self.id       = id
        self.username = username
        self.email    = email


@login_manager.user_loader
def load_user(user_id: str):
    row = _db_one("SELECT id, username, email FROM users WHERE id = ?", (int(user_id),))
    return User(row["id"], row["username"], row["email"] or "") if row else None


# ── SQLite helpers ────────────────────────────────────────────────────────────

def _get_db():
    DB_FILE.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

def _db_exec(sql: str, params: tuple = ()):
    with _get_db() as conn:
        conn.execute(sql, params)

def _db_one(sql: str, params: tuple = ()):
    with _get_db() as conn:
        return conn.execute(sql, params).fetchone()

def _db_all(sql: str, params: tuple = ()):
    with _get_db() as conn:
        return conn.execute(sql, params).fetchall()

def _init_db():
    with _get_db() as conn:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS users (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                username      TEXT UNIQUE NOT NULL,
                email         TEXT UNIQUE,
                password_hash TEXT NOT NULL,
                created_at    DATETIME DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS links (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                keyword    TEXT UNIQUE NOT NULL,
                url        TEXT NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS login_history (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id    INTEGER NOT NULL,
                ip_address TEXT,
                browser    TEXT,
                os         TEXT,
                device     TEXT,
                user_agent TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (user_id) REFERENCES users(id) ON DELETE CASCADE
            );
            CREATE TABLE IF NOT EXISTS chat_groups (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                chat_id       INTEGER UNIQUE NOT NULL,
                title         TEXT NOT NULL DEFAULT '',
                chat_type     TEXT NOT NULL DEFAULT 'unknown',
                role          TEXT NOT NULL DEFAULT 'none',
                is_active     INTEGER NOT NULL DEFAULT 1,
                discovered_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_seen_at  DATETIME DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS app_settings (
                key   TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            INSERT OR IGNORE INTO app_settings (key, value)
                VALUES ('registration_enabled', '1');
        """)
        # Migration: users tablosuna email kolonu
        cols = [r[1] for r in conn.execute("PRAGMA table_info(users)").fetchall()]
        if "email" not in cols:
            conn.execute("ALTER TABLE users ADD COLUMN email TEXT")

    # Migration: links.txt → links tablosu (tek seferlik)
    if _db_one("SELECT COUNT(*) AS n FROM links")["n"] == 0 and LINKS_FILE.exists():
        _import_links_from_file()

def _import_links_from_file():
    """links.txt'deki kayıtları DB'ye aktar (migration)."""
    with _get_db() as conn:
        for line in LINKS_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or ":" not in line:
                continue
            keyword, url = line.split(":", 1)
            conn.execute(
                "INSERT OR IGNORE INTO links (keyword, url) VALUES (?, ?)",
                (keyword.strip(), url.strip()),
            )

def _sync_links_file():
    """Botun okuması için links.txt'yi DB'den yeniden oluşturur."""
    rows = _db_all("SELECT keyword, url FROM links ORDER BY created_at ASC")
    header = (
        "# Format: anahtar_kelime : yeni_link\n"
        "# Anahtar kelime: mesaj metninde VEYA kaynak linkin \"/\" sonrasında aranır\n\n"
    )
    LINKS_FILE.write_text(
        header + "\n".join(f"{r['keyword']} : {r['url']}" for r in rows) + "\n",
        encoding="utf-8",
    )


# ── Ayar helpers ──────────────────────────────────────────────────────────────

def _get_setting(key: str, default: str = "") -> str:
    row = _db_one("SELECT value FROM app_settings WHERE key = ?", (key,))
    return row["value"] if row else default

def _set_setting(key: str, value: str):
    _db_exec("INSERT OR REPLACE INTO app_settings (key, value) VALUES (?, ?)", (key, value))

def _user_count() -> int:
    row = _db_one("SELECT COUNT(*) AS n FROM users")
    return row["n"] if row else 0


def _parse_ua(ua: str) -> tuple[str, str, str]:
    """User-Agent → (browser, os, device)."""
    ua = ua or ""

    if "Edg/" in ua or "Edge/" in ua:
        browser = "Edge"
    elif "OPR/" in ua or "Opera/" in ua:
        browser = "Opera"
    elif "Chrome/" in ua:
        browser = "Chrome"
    elif "Firefox/" in ua:
        browser = "Firefox"
    elif "Safari/" in ua:
        browser = "Safari"
    else:
        browser = "Bilinmiyor"

    if "Windows NT" in ua:
        os_name = "Windows"
    elif "Macintosh" in ua:
        os_name = "macOS"
    elif "Android" in ua:
        os_name = "Android"
    elif "iPhone" in ua or "iPad" in ua:
        os_name = "iOS"
    elif "Linux" in ua:
        os_name = "Linux"
    else:
        os_name = "Bilinmiyor"

    if "iPhone" in ua or ("Android" in ua and "Mobile" in ua):
        device = "Mobil"
    elif "iPad" in ua or ("Android" in ua and "Mobile" not in ua):
        device = "Tablet"
    else:
        device = "Masaüstü"

    return browser, os_name, device


def _record_login(user_id: int):
    ua = request.headers.get("User-Agent", "")
    browser, os_name, device = _parse_ua(ua)
    ip = request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip()
    _db_exec(
        "INSERT INTO login_history (user_id, ip_address, browser, os, device, user_agent)"
        " VALUES (?, ?, ?, ?, ?, ?)",
        (user_id, ip, browser, os_name, device, ua[:500]),
    )


# ── .env helpers ──────────────────────────────────────────────────────────────

def load_env() -> dict:
    config = {}
    if not ENV_FILE.exists():
        return config
    for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, val = line.split("=", 1)
        config[key.strip()] = val.strip()
    return config

def save_env(config: dict):
    ENV_FILE.write_text("".join(f"{k}={v}\n" for k, v in config.items()), encoding="utf-8")


# ── Auth routes ───────────────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))
    if request.method == "POST":
        email    = request.form.get("email",    "").strip().lower()
        password = request.form.get("password", "")
        row = _db_one(
            "SELECT id, username, email, password_hash FROM users WHERE email = ?", (email,)
        )
        if row and check_password_hash(row["password_hash"], password):
            login_user(User(row["id"], row["username"], row["email"] or ""),
                       remember=bool(request.form.get("remember")))
            _record_login(row["id"])
            return redirect(request.args.get("next") or url_for("index"))
        flash("E-posta veya şifre hatalı.", "danger")
    return render_template("login.html")


@app.route("/register", methods=["GET", "POST"])
def register():
    count    = _user_count()
    reg_open = _get_setting("registration_enabled", "1") == "1"
    if not reg_open and count > 0:
        flash("Kayıt şu anda devre dışı.", "danger")
        return redirect(url_for("login"))
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        username  = request.form.get("username",  "").strip()
        email     = request.form.get("email",     "").strip().lower()
        password  = request.form.get("password",  "")
        password2 = request.form.get("password2", "")
        if not username or not email or not password:
            flash("Tüm alanlar zorunludur.", "danger")
        elif "@" not in email or "." not in email.split("@")[-1]:
            flash("Geçerli bir e-posta adresi girin.", "danger")
        elif len(password) < 6:
            flash("Şifre en az 6 karakter olmalıdır.", "danger")
        elif password != password2:
            flash("Şifreler eşleşmiyor.", "danger")
        else:
            try:
                _db_exec(
                    "INSERT INTO users (username, email, password_hash) VALUES (?, ?, ?)",
                    (username, email, generate_password_hash(password)),
                )
                flash("Kayıt başarılı! Giriş yapabilirsiniz.", "success")
                return redirect(url_for("login"))
            except sqlite3.IntegrityError as e:
                flash("Bu kullanıcı adı zaten alınmış." if "username" in str(e)
                      else "Bu e-posta adresi zaten kayıtlı.", "warning")

    return render_template("register.html", first_setup=(count == 0))


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


# ── HTML sayfaları (sadece template render eder) ──────────────────────────────

@app.route("/")
@login_required
def index():
    links = [dict(r) for r in _db_all(
        "SELECT keyword, url, created_at FROM links ORDER BY created_at DESC LIMIT 5"
    )]
    _row = _db_one("SELECT COUNT(*) AS n FROM links")
    total_links = _row["n"] if _row else 0
    sources = [dict(r) for r in _db_all(
        "SELECT chat_id, title FROM chat_groups WHERE role='source' AND is_active=1 ORDER BY title"
    )]
    targets = [dict(r) for r in _db_all(
        "SELECT chat_id, title FROM chat_groups WHERE role='target' AND is_active=1 ORDER BY title"
    )]
    return render_template("index.html",
                           links=links,
                           total_links=total_links,
                           sources=sources,
                           targets=targets)

@app.route("/links")
@login_required
def links_page():
    return render_template("links.html")

@app.route("/profile")
@login_required
def profile():
    return render_template("profile.html")

@app.route("/settings")
@login_required
def settings():
    return render_template("settings.html")

@app.route("/groups")
@login_required
def groups_page():
    return render_template("groups.html")


# ── JSON API ──────────────────────────────────────────────────────────────────

def _api_err(msg: str, code: int = 400):
    return jsonify({"error": msg}), code

def _require_json():
    """Returns parsed JSON or None."""
    return request.get_json(silent=True) or {}


# Links API
@app.route("/api/links")
@login_required
def api_links():
    rows = _db_all("SELECT keyword, url, created_at FROM links ORDER BY created_at DESC")
    return jsonify([dict(r) for r in rows])

@app.route("/api/links", methods=["POST"])
@login_required
def api_add_link():
    d       = _require_json()
    keyword = d.get("keyword", "").strip()
    url     = d.get("url",     "").strip()
    if not keyword or not url:
        return _api_err("Anahtar kelime ve URL boş olamaz.")
    try:
        _db_exec("INSERT INTO links (keyword, url) VALUES (?, ?)", (keyword, url))
        _sync_links_file()
        row = _db_one("SELECT keyword, url, created_at FROM links WHERE keyword = ?", (keyword,))
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return _api_err(f"'{keyword}' zaten mevcut.", 409)

@app.route("/api/links/<path:keyword>", methods=["PUT"])
@login_required
def api_edit_link(keyword):
    d       = _require_json()
    new_kw  = d.get("keyword", "").strip()
    new_url = d.get("url",     "").strip()
    if not new_kw or not new_url:
        return _api_err("Anahtar kelime ve URL boş olamaz.")
    try:
        _db_exec(
            "UPDATE links SET keyword = ?, url = ? WHERE keyword = ?",
            (new_kw, new_url, keyword),
        )
        _sync_links_file()
        row = _db_one("SELECT keyword, url, created_at FROM links WHERE keyword = ?", (new_kw,))
        return jsonify(dict(row))
    except sqlite3.IntegrityError:
        return _api_err(f"'{new_kw}' zaten mevcut.", 409)

@app.route("/api/links/<path:keyword>", methods=["DELETE"])
@login_required
def api_delete_link(keyword):
    _db_exec("DELETE FROM links WHERE keyword = ?", (keyword,))
    _sync_links_file()
    return jsonify({"ok": True})


# Groups API
@app.route("/api/groups")
@login_required
def api_groups():
    rows = _db_all(
        "SELECT * FROM chat_groups ORDER BY"
        " CASE role WHEN 'source' THEN 0 WHEN 'target' THEN 1 ELSE 2 END,"
        " title, chat_id"
    )
    return jsonify([dict(r) for r in rows])


@app.route("/api/groups", methods=["POST"])
@login_required
def api_group_add():
    d = _require_json()
    try:
        chat_id = int(d.get("chat_id", 0))
    except (ValueError, TypeError):
        return _api_err("Geçersiz chat_id.")
    if not chat_id:
        return _api_err("chat_id zorunludur.")
    title = d.get("title", "").strip()
    role  = d.get("role",  "none")
    if role not in ("source", "target", "none"):
        return _api_err("Geçersiz rol.")
    try:
        _db_exec(
            "INSERT INTO chat_groups (chat_id, title, role) VALUES (?, ?, ?)",
            (chat_id, title, role),
        )
        row = _db_one("SELECT * FROM chat_groups WHERE chat_id = ?", (chat_id,))
        return jsonify(dict(row)), 201
    except sqlite3.IntegrityError:
        return _api_err("Bu chat_id zaten kayıtlı.", 409)


@app.route("/api/groups/<int:gid>/role", methods=["POST"])
@login_required
def api_group_role(gid):
    d    = _require_json()
    role = d.get("role", "none")
    if role not in ("source", "target", "none"):
        return _api_err("Geçersiz rol.")
    _db_exec("UPDATE chat_groups SET role = ? WHERE id = ?", (role, gid))
    return jsonify({"ok": True})


@app.route("/api/groups/<int:gid>/toggle", methods=["POST"])
@login_required
def api_group_toggle(gid):
    _db_exec("UPDATE chat_groups SET is_active = 1 - is_active WHERE id = ?", (gid,))
    row = _db_one("SELECT is_active FROM chat_groups WHERE id = ?", (gid,))
    return jsonify({"is_active": row["is_active"] if row else 0})


@app.route("/api/groups/<int:gid>", methods=["PATCH"])
@login_required
def api_group_edit(gid):
    d = _require_json()
    title = d.get("title", "").strip()
    try:
        chat_id = int(d.get("chat_id", 0))
    except (ValueError, TypeError):
        return _api_err("Geçersiz chat_id.")
    if not chat_id:
        return _api_err("chat_id zorunludur.")
    try:
        _db_exec(
            "UPDATE chat_groups SET title = ?, chat_id = ? WHERE id = ?",
            (title, chat_id, gid),
        )
        row = _db_one("SELECT * FROM chat_groups WHERE id = ?", (gid,))
        return jsonify(dict(row))
    except sqlite3.IntegrityError:
        return _api_err("Bu chat_id zaten kayıtlı.", 409)


@app.route("/api/groups/<int:gid>", methods=["DELETE"])
@login_required
def api_group_delete(gid):
    _db_exec("DELETE FROM chat_groups WHERE id = ?", (gid,))
    return jsonify({"ok": True})


# Settings API
@app.route("/api/settings")
@login_required
def api_settings_get():
    cfg = load_env()
    return jsonify({
        "BOT_TOKEN":            cfg.get("BOT_TOKEN", ""),
        "APP_URL":              cfg.get("APP_URL",   APP_URL),
        "APP_PORT":             cfg.get("APP_PORT",  str(APP_PORT)),
        "registration_enabled": _get_setting("registration_enabled", "1") == "1",
        "user_count":           _user_count(),
    })

@app.route("/api/settings/bot", methods=["POST"])
@login_required
def api_settings_bot():
    d = _require_json()
    if not d.get("BOT_TOKEN"):
        return _api_err("Bot Token zorunludur.")
    cfg = load_env()
    cfg["BOT_TOKEN"] = d.get("BOT_TOKEN", "").strip()
    if d.get("APP_URL"):
        cfg["APP_URL"] = d.get("APP_URL", "").strip().rstrip("/")
    if d.get("APP_PORT"):
        try:
            cfg["APP_PORT"] = str(int(d.get("APP_PORT")))
        except (ValueError, TypeError):
            return _api_err("APP_PORT geçerli bir sayı olmalıdır.")
    save_env(cfg)
    return jsonify({"ok": True})

@app.route("/api/settings/app", methods=["POST"])
@login_required
def api_settings_app():
    d = _require_json()
    _set_setting("registration_enabled", "1" if d.get("registration_enabled") else "0")
    return jsonify({"ok": True})


# Profile API
@app.route("/api/profile")
@login_required
def api_profile_get():
    row = _db_one("SELECT username, email FROM users WHERE id = ?", (current_user.id,))
    return jsonify(dict(row))

@app.route("/api/profile/info", methods=["POST"])
@login_required
def api_profile_info():
    d        = _require_json()
    username = d.get("username", "").strip()
    email    = d.get("email",    "").strip().lower()
    if not username or not email:
        return _api_err("Kullanıcı adı ve e-posta boş olamaz.")
    if "@" not in email or "." not in email.split("@")[-1]:
        return _api_err("Geçerli bir e-posta adresi girin.")
    try:
        _db_exec(
            "UPDATE users SET username = ?, email = ? WHERE id = ?",
            (username, email, current_user.id),
        )
        return jsonify({"ok": True, "username": username, "email": email})
    except sqlite3.IntegrityError:
        return _api_err("Bu kullanıcı adı veya e-posta zaten alınmış.", 409)

@app.route("/api/profile/password", methods=["POST"])
@login_required
def api_profile_password():
    d          = _require_json()
    current_pw = d.get("current_password", "")
    new_pw     = d.get("new_password",     "")
    confirm_pw = d.get("confirm_password", "")
    row = _db_one("SELECT password_hash FROM users WHERE id = ?", (current_user.id,))
    if not check_password_hash(row["password_hash"], current_pw):
        return _api_err("Mevcut şifre hatalı.")
    if len(new_pw) < 6:
        return _api_err("Yeni şifre en az 6 karakter olmalıdır.")
    if new_pw != confirm_pw:
        return _api_err("Şifreler eşleşmiyor.")
    _db_exec(
        "UPDATE users SET password_hash = ? WHERE id = ?",
        (generate_password_hash(new_pw), current_user.id),
    )
    return jsonify({"ok": True})

@app.route("/api/profile/history")
@login_required
def api_profile_history():
    rows = _db_all(
        "SELECT ip_address, browser, os, device, created_at"
        " FROM login_history WHERE user_id = ?"
        " ORDER BY created_at DESC",
        (current_user.id,),
    )
    return jsonify([dict(r) for r in rows])


# ── Export / Import ──────────────────────────────────────────────────────────

@app.route("/api/links/export/txt")
@login_required
def api_export_txt():
    rows = _db_all("SELECT keyword, url FROM links ORDER BY created_at ASC")
    body = "\n".join(f"{r['keyword']} : {r['url']}" for r in rows)
    resp = make_response(body)
    resp.headers["Content-Type"]        = "text/plain; charset=utf-8"
    resp.headers["Content-Disposition"] = "attachment; filename=links.txt"
    return resp


@app.route("/api/links/export/csv")
@login_required
def api_export_csv():
    rows = _db_all("SELECT keyword, url, created_at FROM links ORDER BY created_at ASC")
    buf  = io.StringIO()
    w    = csv_mod.writer(buf)
    w.writerow(["keyword", "url", "created_at"])
    for r in rows:
        w.writerow([r["keyword"], r["url"], r["created_at"] or ""])
    resp = make_response(buf.getvalue())
    resp.headers["Content-Type"]        = "text/csv; charset=utf-8"
    resp.headers["Content-Disposition"] = "attachment; filename=links.csv"
    return resp


@app.route("/api/links/export/excel")
@login_required
def api_export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    rows = _db_all("SELECT keyword, url, created_at FROM links ORDER BY created_at ASC")
    wb   = Workbook()
    ws   = wb.active
    ws.title = "Linkler"
    ws.append(["Anahtar Kelime", "Hedef URL", "Eklenme Tarihi"])
    for cell in ws[1]:
        cell.font      = Font(bold=True)
        cell.fill      = PatternFill("solid", fgColor="4F46E5")
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 22
    for r in rows:
        ws.append([r["keyword"], r["url"], r["created_at"] or ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers["Content-Type"]        = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] = "attachment; filename=links.xlsx"
    return resp


@app.route("/api/links/import", methods=["POST"])
@login_required
def api_import_links():
    if "file" not in request.files or not request.files["file"].filename:
        return _api_err("Dosya bulunamadı.")
    file     = request.files["file"]
    fname    = file.filename.lower()
    pairs: list[tuple[str, str]] = []

    try:
        if fname.endswith(".txt"):
            for line in file.read().decode("utf-8-sig").splitlines():
                line = line.strip()
                if not line or line.startswith("#"):
                    continue
                if " : " in line:
                    kw, url = line.split(" : ", 1)
                elif ":" in line:
                    # fallback: split on first colon but skip if it looks like http://
                    parts = line.split(":", 1)
                    kw, url = parts[0], parts[1]
                else:
                    continue
                if kw.strip() and url.strip():
                    pairs.append((kw.strip(), url.strip()))

        elif fname.endswith(".csv"):
            text   = file.read().decode("utf-8-sig")
            reader = list(csv_mod.reader(io.StringIO(text)))
            start  = 1 if reader and reader[0][0].lower() in ("keyword","anahtar kelime","anahtar_kelime") else 0
            for row in reader[start:]:
                if len(row) >= 2 and row[0].strip() and row[1].strip():
                    pairs.append((row[0].strip(), row[1].strip()))

        elif fname.endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file.read()))
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0 and row[0] and str(row[0]).lower() in ("keyword","anahtar kelime","anahtar_kelime"):
                    continue
                if row[0] and row[1]:
                    pairs.append((str(row[0]).strip(), str(row[1]).strip()))
        else:
            return _api_err("Desteklenmeyen format. .txt, .csv veya .xlsx yükleyin.")
    except Exception as e:
        return _api_err(f"Dosya okunamadı: {e}")

    added = skipped = 0
    with _get_db() as conn:
        for kw, url in pairs:
            if not kw or not url:
                continue
            try:
                conn.execute("INSERT INTO links (keyword, url) VALUES (?, ?)", (kw, url))
                added += 1
            except sqlite3.IntegrityError:
                skipped += 1
    if added:
        _sync_links_file()
    return jsonify({"added": added, "skipped": skipped, "total": len(pairs)})


# ── Startup ───────────────────────────────────────────────────────────────────

with app.app_context():
    _init_db()

if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=APP_PORT,
        debug=os.environ.get("FLASK_DEBUG", "false").lower() == "true",
    )
